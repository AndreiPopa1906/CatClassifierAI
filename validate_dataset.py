import openpyxl
# Load the workbook and select a worksheet
path_database = r'C:\Users\matei\Documents\AI_project\MModified_cat_database.xlsx'
path_new_database=r'C:\Users\matei\Documents\AI_project\RE7_modified_cat_database.xlsx'
workbook = openpyxl.load_workbook(path_database)

# Save all changes once
def init():
    sheet = workbook['Code']
    sheet['B5'] = "1/2/3/4/5"  # Modify the value in B5 of 'Code' sheet
    workbook.save(path_database)  # Save changes after both modifications




def delete_rows(to_delete_rows, sheet):
    # Sort rows in descending order to avoid index issues when deleting
    to_delete_rows = sorted(to_delete_rows, reverse=True)
    print("Avem de sters :",len(to_delete_rows))
    # Delete rows directly
    for row in to_delete_rows:
        sheet.delete_rows(row)

    # print(f"Deleted rows: {to_delete_rows}")


def get_categories():
    sheet = workbook['Code']
    map_structure = {}
    row_index = 2
    while True:
        cell_value = sheet[f'A{row_index}'].value  # Get the value of the cell in column A
        if cell_value is None:  # Check if the cell is blank
            break  # Exit the loop if a blank cell is found
        possible_values = get_possible_values(row_index)
        map_structure[cell_value] = possible_values
        row_index += 1  # Move to the next row
    return map_structure

def get_possible_values(row_index):
    sheet = workbook['Code']
    cell_value = sheet[f'B{row_index}'].value  # Get the value of the cell in column B
    if cell_value is None:
        return []  # Return an empty list if the cell is blank
    possible_values = list(map(str.strip, cell_value.split('/')))  # Use str.strip() to remove extra spaces
    return possible_values

# New function to search for row duplicates based on values from columns C to AB
def search_for_row_duplicates():
    sheet = workbook['Data']
    seen_rows = {}  # Dictionary to store combinations of values and their row index
    duplicate_rows = []  # List to store the row numbers of duplicates

    # Iterate through all rows in the sheet starting from row 2
    for row_index in range(2, sheet.max_row + 1):
        # Extract values from columns C to AB (3 to 28)
        row_values = tuple(sheet.cell(row=row_index, column=col_index).value for col_index in range(3, 29))

        if row_values in seen_rows:
            # If the combination has been seen, it's a duplicate
            duplicate_rows.append(row_index)
            print(f"Duplicate found in row {row_index}, matches row {seen_rows[row_values]}")
        else:
            # Store the row's values as a key in the dictionary
            seen_rows[row_values] = row_index

    # if duplicate_rows:
    #     print(f"Duplicate rows found: {duplicate_rows}")
    # else:
    #     print("No duplicate rows found.")
    return duplicate_rows

def validate_data():
    sheet = workbook['Data']
    map_structure = get_categories()  # Fetch valid categories
    invalid_rows = []  # List to store rows with invalid data

    # Get the column names from the first row (row=1)
    column_names = {col_index: sheet.cell(row=1, column=col_index).value for col_index in range(3, 29)}
    error_mapped = 0
    error_unmapped = 0
    invalid_rows_number = 0
    invalid_data_category_set = set()
    to_delete_rows=set()
    # Iterate through all rows in the sheet, starting at row 2 (assuming row 1 is the header)
    for row_index in range(2, sheet.max_row + 1):
        row_invalid = False  # Track if the row contains any invalid data

        # Iterate through columns C to AB (columns 3 to 28)
        for col_index in range(3, 29):
            cell_value = sheet.cell(row=row_index, column=col_index).value
            column_name = column_names[col_index]

            if column_name in map_structure:
                # Get possible values for this column
                valid_values = map_structure[column_name]
                if isinstance(cell_value, int):
                    cell_value = str(cell_value)

                if cell_value not in valid_values:
                    row_invalid = True
                    error_mapped += 1
                    invalid_data_category_set.add((column_name, cell_value))
                    if(column_name=="Race"):
                        to_delete_rows.add(row_index)

            else:
                # For columns not in map_structure, the value must be between 1 and 5
                try:
                    cell_value = int(cell_value)
                    if not (1 <= cell_value <= 5):
                        row_invalid = True
                        error_unmapped += 1
                        invalid_data_category_set.add((column_name, cell_value))
                except (ValueError, TypeError):
                    row_invalid = True  # Mark the row as invalid

        if row_invalid:
            invalid_rows_number += 1  # Increment the count for invalid rows

    duplicates=search_for_row_duplicates()
    for rows_index in duplicates:
        to_delete_rows.add(rows_index)
    delete_rows(to_delete_rows,workbook["Data"])
    workbook.save(path_new_database)
    for element in invalid_data_category_set:
        print(element)

# Initialize and fetch category mappings
init()
# Validate data in the "Data" sheet
validate_data()



