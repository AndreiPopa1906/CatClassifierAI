import openpyxl

# Load the workbook and select a worksheet
path_database = r'C:\Users\matei\Documents\AI_project\cat_database.xlsx'
workbook = openpyxl.load_workbook(path_database)  # Replace with your file name
path_modified_database=r'C:\Users\matei\Documents\AI_project\MModified_cat_database.xlsx'

def return_column_index(column_header):
    sheet = workbook['Data']
    row_index = 1  # Row 1, assuming column headers are in the first row
    column_index = 1  # Start from the first column
    cell_value = sheet.cell(row=row_index, column=column_index).value

    # Log the column header search


    while cell_value != column_header:
        column_index += 1  # Move to the next column
        cell_value = sheet.cell(row=row_index, column=column_index).value

        if cell_value is None:  # If no header found, stop
            raise ValueError(f"Column '{column_header}' not found.")

        # Log progress during search



    return column_index


def modify_values_column_ext():
    sheet = workbook['Data']
    column_index = return_column_index("Ext")

    print(f"Modifying values in the 'Ext' column (column {column_index})...")

    for i in range(2, sheet.max_row+1):  # Assuming rows 2 to 9 should be updated
        cell_value = sheet.cell(row=i, column=column_index).value
        print(f"Original value in row {i}, column {column_index}: {cell_value}")

        if cell_value is None:  # Handle empty cells
            print(f"Skipping row {i}, empty cell.")
            continue

        try:
            aux = int(cell_value)  # Convert cell value to integer
            aux += 1  # Increment the value
            sheet.cell(row=i, column=column_index).value = aux  # Update the same cell
            print(f"Updated value in row {i}, column {column_index}: {aux}")
        except ValueError:
            print(f"Warning: Cell at row {i}, column {column_index} contains non-integer value '{cell_value}'")

    workbook.save(path_modified_database)  # Save all changes once
    print(f"Workbook saved to {path_database}")

def modify_values_column_obs():
    sheet = workbook['Data']
    column_index = return_column_index("Obs")

    for i in range(2, sheet.max_row+1):
        cell_value = sheet.cell(row=i, column=column_index).value
        if cell_value is None:  # Handle empty cells
            continue

        try:
            aux = int(cell_value)  # Convert cell value to integer
            aux += 1  # Increment the value
            sheet.cell(row=i, column=column_index).value = aux  # Update the same cell
        except ValueError:
            print(f"Warning: Cell at row {i}, column {column_index} is not an integer.")

def modify_values_column_pred_mamm():
    sheet = workbook['Data']
    column_index = return_column_index("PredMamm")

    for i in range(2, sheet.max_row+1):
        cell_value = sheet.cell(row=i, column=column_index).value
        if cell_value is None:  # Handle empty cells
            continue

        try:
            aux = int(cell_value)  # Convert cell value to integer
            aux += 1  # Increment the value
            sheet.cell(row=i, column=column_index).value = aux  # Update the same cell
        except ValueError:
            print(f"Warning: Cell at row {i}, column {column_index} is not an integer.")

def modify_values_column_pred_oiseau():
    sheet = workbook['Data']
    column_index = return_column_index("PredOiseau")

    for i in range(2, sheet.max_row+1):
        cell_value = sheet.cell(row=i, column=column_index).value
        if cell_value is None:  # Handle empty cells
            continue

        try:
            aux = int(cell_value)  # Convert cell value to integer
            aux += 1  # Increment the value
            sheet.cell(row=i, column=column_index).value = aux  # Update the same cell
        except ValueError:
            print(f"Warning: Cell at row {i}, column {column_index} is not an integer.")


def modify_values_column_nombre():
    sheet = workbook['Data']
    column_index = return_column_index("Nombre")

    for i in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=i, column=column_index).value
        if(cell_value=="Plusde5"):
            sheet.cell(row=i, column=column_index).value=str(5)
def init():
    sheet = workbook['Code']
    sheet['B5'] = "1/2/3/4/5"  # Modify the value in B5 of 'Code' sheet
    print(f"Updated 'Code' sheet cell B5: 1/2/3/4/5")
    modify_values_column_ext()
    modify_values_column_obs()
    modify_values_column_pred_mamm()
    modify_values_column_pred_oiseau()
    modify_values_column_nombre()
    workbook.save(path_modified_database)

    print("All changes saved successfully.")


# Call the init function to run the modifications
init()
